# -*- coding: utf-8 -*-
__author__ = 'tyr'
import openpyxl


class OpenrateExcel(object):
    def __init__(self, path, sheet_name):
        self.wb = openpyxl.load_workbook(path)
        self.sheet = self.wb[sheet_name]

    def get_all_sheets(self):
        return self.wb.sheetnames

    def get_max_row(self):
        return self.sheet.max_row

    def get_max_column(self):
        return self.sheet.max_column

    def get_value(self, row, column):
        return self.sheet.cell(row, column).value

    def get_values(self, row_start, row_count, column_start, column_count):
        values = []
        for i in range(row_start, row_count + 1):
            for j in range(column_start, column_count + 1):
                values.append(self.sheet.cell(row=i, column=j).value)
        return values

    def set_value(self, row, column, value):
        self.sheet.cell(row=row, column=column).value = value

    def save(self, path):
        self.wb.save(path)
